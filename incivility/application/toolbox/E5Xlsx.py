import dataclasses
import os
from enum import Enum
from pathlib import Path
from typing import ClassVar, Union, Tuple, Optional

import openpyxl


# E5
class SheetNameEnum(Enum):
    INCIVILITY = "Incivility"
    DELAY = "Delay"
    ABSENCE = "Absence"


# E5
@dataclasses.dataclass
class E5Xlsx:
    # Class Variables
    INCIVILITY_HEADER: ClassVar[list[str]] = ["Date", "Nom de l'élève", "N° et type d'incivilité", "Détail"]
    DELAY_HEADER: ClassVar[list[str]] = ["Date", "Nom de l'élève", "Justifié", "Détail"]
    ABSENCE_HEADER: ClassVar[list[str]] = ["Date", "Nom de l'élève", "Durée", "Justifiée", "Détail"]
    BASE_DIR: ClassVar[Path] = Path(__file__).resolve().parent.parent
    INCIVILITYS_DIRECTORY: ClassVar[Path] = os.path.join(BASE_DIR / "incivility_csv")

    # Instance Variables
    workbook: Union[openpyxl.Workbook, None] = None
    full_path_csv_file: str = ""

    # E5
    def check_if_file_exists(self) -> tuple[bool, str, Optional[bool]]:
        success: bool = True
        message: str = ""
        file_exists: Union[bool, None] = None

        try:
            file_exists = os.path.exists(self.full_path_csv_file)
        except Exception as ex:
            success = False
            message = f"E5Xlsx.check_if_file_exists() Error : {ex}"

        return success, message, file_exists

    # E5
    def append_to_sheet(self, sheet_name: str, data: list[str]) -> (bool, str):
        success: bool = True
        message: str = ""

        try:
            self.workbook = openpyxl.load_workbook(filename=self.full_path_csv_file)
            target_sheet = self.workbook[sheet_name]
            target_sheet.append(data)
            self.workbook.save(filename=self.full_path_csv_file)
        except Exception as ex:
            success = False
            message = f"E5Xlsx.append_to_sheet() Error : {ex}"

        return success, message

    # E5
    def create_workbook(self) -> (bool, str):
        success: bool = True
        message: str = ""

        try:
            self.workbook = openpyxl.Workbook()

            # Create Sheets
            incivility_sheet = self.workbook.active
            incivility_sheet.title = SheetNameEnum.INCIVILITY.value
            self.workbook.create_sheet(SheetNameEnum.DELAY.value)
            self.workbook.create_sheet(SheetNameEnum.ABSENCE.value)

            # Append Headers
            incivility_sheet.append(self.INCIVILITY_HEADER)
            self.workbook[SheetNameEnum.DELAY.value].append(self.DELAY_HEADER)
            self.workbook[SheetNameEnum.ABSENCE.value].append(self.ABSENCE_HEADER)

            # Save Workbook
            self.workbook.save(filename=self.full_path_csv_file)
        except Exception as ex:
            success = False
            message = f"E5Xlsx.create_workbook() Error : {ex}"

        return success, message
